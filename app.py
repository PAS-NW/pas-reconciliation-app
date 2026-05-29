import io
import re
import zipfile
from pathlib import Path
from datetime import datetime
from urllib.parse import quote
from html import escape
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

try:
    from pypdf import PdfReader
except Exception:
    try:
        from PyPDF2 import PdfReader
    except Exception:
        PdfReader = None

PAS_YELLOW = "#FFD400"
PAS_BLACK = "#0A0A0A"
PAS_DARK = "#171717"
PAS_GREY = "#F4F4F4"

st.set_page_config(page_title="PAS Plant Invoice Matching", page_icon="pas_logo.png", layout="wide")

st.markdown(
    f"""
    <style>
    .stApp {{ background: #f5f5f5; color: #0A0A0A; }}
    section[data-testid="stSidebar"] {{
        background: {PAS_BLACK};
        color: white;
        padding-top: 1.45rem;
    }}
    section[data-testid="stSidebar"] * {{ color: white; }}
    section[data-testid="stSidebar"] img {{
        margin-top: 0.15rem;
        border-radius: 14px;
    }}
    .block-container {{
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1500px;
    }}

    .pas-hero {{
        background: linear-gradient(135deg, {PAS_BLACK} 0%, #202020 70%, #7a6900 135%);
        border-radius: 18px;
        padding: 24px 28px;
        margin-bottom: 18px;
        box-shadow: 0 8px 25px rgba(0,0,0,0.12);
    }}
    .pas-title {{
        color: white;
        font-size: 32px;
        font-weight: 900;
        margin: 0;
        letter-spacing: -0.03em;
    }}
    .pas-subtitle {{
        color: {PAS_YELLOW};
        font-size: 14px;
        margin-top: 4px;
        font-weight: 800;
    }}

    .kpi-card {{
        background: white;
        border-radius: 18px;
        padding: 18px 20px;
        border: 1px solid #e8e8e8;
        box-shadow: 0 3px 12px rgba(0,0,0,0.05);
        min-height: 112px;
    }}
    .kpi-label {{
        color: #111;
        font-size: 14px;
        font-weight: 800;
        margin-bottom: 8px;
    }}
    .kpi-value {{
        color: {PAS_YELLOW};
        font-size: 36px;
        font-weight: 950;
        line-height: 1.05;
        text-shadow: 0 1px 0 #111;
    }}
    .kpi-sub {{
        color: #222;
        font-size: 13px;
        margin-top: 6px;
    }}

    .stButton > button, .stDownloadButton > button {{
        background: {PAS_YELLOW} !important;
        color: {PAS_BLACK} !important;
        border: 1px solid {PAS_BLACK} !important;
        border-radius: 12px !important;
        font-weight: 900 !important;
    }}

    /* Keep app helper text readable */
    .stCaption, div[data-testid="stCaptionContainer"], .stMarkdown p, .stInfo {{
        color: #0A0A0A !important;
    }}

    .pas-results-title {{
        color: #0A0A0A;
        font-size: 26px;
        font-weight: 950;
        margin: 22px 0 8px 0;
    }}
    .pas-unmatched-pill {{
        display: inline-flex;
        align-items: center;
        gap: 8px;
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        border: 1px solid #111;
        border-radius: 14px 14px 0 0;
        padding: 11px 18px;
        font-weight: 950;
        box-shadow: 0 3px 10px rgba(0,0,0,0.08);
        margin-top: 4px;
    }}

    .pas-table-wrap {{
        background: white;
        border: 1px solid #d9d9d9;
        border-radius: 0 16px 16px 16px;
        overflow: auto;
        box-shadow: 0 4px 18px rgba(0,0,0,0.07);
        margin-bottom: 18px;
    }}
    table.pas-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 13px;
        color: #0A0A0A;
        background: white;
    }}
    table.pas-table thead th {{
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        font-weight: 950;
        text-align: left;
        padding: 11px 12px;
        border: 1px solid #c7a900;
        white-space: nowrap;
    }}
    table.pas-table tbody td {{
        background: white;
        color: #0A0A0A;
        padding: 9px 12px;
        border: 1px solid #e3e3e3;
        vertical-align: top;
    }}
    table.pas-table tbody tr:nth-child(even) td {{
        background: #fbfbfb;
    }}
    table.pas-table a {{
        color: #006fd6 !important;
        font-weight: 800;
        text-decoration: none;
    }}
    table.pas-table a:hover {{
        text-decoration: underline;
    }}
    .pas-note {{
        color: #0A0A0A;
        font-size: 13px;
        margin: 8px 0 16px 0;
    }}
    .pas-support {{
        color: #0A0A0A;
        font-size: 14px;
        margin: 16px 0;
    }}
    .pas-support a {{
        color: #006fd6 !important;
        font-weight: 800;
    }}


    /* --- hard hide Streamlit's uploaded-file chip/list while keeping uploader button usable --- */
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"],
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFileName"],
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFileSize"],
    div[data-testid="stFileUploader"] ul,
    div[data-testid="stFileUploader"] div[role="list"],
    div[data-testid="stFileUploader"] div[role="listitem"] {{
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
        min-height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
        overflow: hidden !important;
    }}
    div[data-testid="stFileUploader"] div:has(button[title*="Remove"]),
    div[data-testid="stFileUploader"] div:has(button[aria-label*="Remove"]),
    div[data-testid="stFileUploader"] div:has(svg[data-testid="DeleteIcon"]) {{
        display: none !important;
    }}
    div[data-testid="stFileUploader"] section > div:not(:has(button)) {{
        display: none !important;
    }}
    div[data-testid="stFileUploader"] button {{
        background: #ffffff !important;
        color: #0A0A0A !important;
        border: 1px solid #d7dce3 !important;
        border-radius: 10px !important;
        font-weight: 900 !important;
        box-shadow: 0 2px 8px rgba(0,0,0,.06) !important;
    }}
    div[data-testid="stFileUploader"] button * {{ color:#0A0A0A !important; fill:#0A0A0A !important; stroke:#0A0A0A !important; }}

    </style>
    """,
    unsafe_allow_html=True,
)


st.markdown(
    """
    <style>
    /* Keep sidebar readable on black */
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown li,
    section[data-testid="stSidebar"] .stMarkdown h1,
    section[data-testid="stSidebar"] .stMarkdown h2,
    section[data-testid="stSidebar"] .stMarkdown h3,
    section[data-testid="stSidebar"] .stMarkdown strong,
    section[data-testid="stSidebar"] .stMarkdown span {
        color: #ffffff !important;
    }

    /* Make upload icons visible on dark bars */
    div[data-testid="stFileUploader"] svg,
    div[data-testid="stFileUploader"] button svg,
    div[data-testid="stFileUploader"] [data-testid="stIconMaterial"] {
        color: #FFD400 !important;
        fill: #FFD400 !important;
        stroke: #FFD400 !important;
    }
    div[data-testid="stFileUploader"] section {
        background: #24242d !important;
        border: 1px solid #30303a !important;
        border-radius: 12px !important;
    }
    div[data-testid="stFileUploader"] button {
        color: white !important;
        border-color: #454552 !important;
        background: #111217 !important;
    }

    /* Results table: white body, yellow sticky header, 10-row scroll area */
    .pas-table-wrap {
        max-height: 510px !important;
        overflow-y: auto !important;
        overflow-x: auto !important;
    }
    .pas-table-wrap thead th {
        position: sticky;
        top: 0;
        z-index: 2;
    }
    .pas-note, .pas-support, .pas-support * {
        color: #0A0A0A !important;
    }

    /* Bottom chase animation: small, low, runs once */
    .pas-bottom-chase-wrap {
        position: fixed;
        left: calc(18rem + 22px);
        right: 42px;
        bottom: 12px;
        height: 58px;
        pointer-events: none;
        z-index: 1;
        overflow: hidden;
    }
    .pas-bottom-ground {
        position: absolute;
        left: 0;
        right: 0;
        bottom: 6px;
        border-bottom: 1px solid rgba(0,0,0,0.11);
    }
    .pas-chase-pack {
        position: absolute;
        bottom: 8px;
        left: -150px;
        width: 150px;
        height: 48px;
        animation: pas-chase-run 13s linear 1 forwards;
    }
    @keyframes pas-chase-run {
        0% { transform: translateX(-120px); opacity: 0; }
        8% { opacity: 1; }
        88% { opacity: 1; }
        100% { transform: translateX(calc(100vw - 90px)); opacity: 0; }
    }
    .pas-truck-mini {
        position: absolute;
        left: 0;
        bottom: 5px;
        width: 54px;
        height: 30px;
        filter: drop-shadow(0 1px 1px rgba(0,0,0,.22));
    }
    .pas-truck-bed {
        position: absolute;
        left: 0;
        top: 5px;
        width: 34px;
        height: 19px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-radius: 4px 2px 3px 5px;
        transform: skewX(-10deg);
    }
    .pas-truck-logo {
        position: absolute;
        left: 7px;
        top: 9px;
        font-size: 9px;
        font-weight: 950;
        color: #0A0A0A;
        line-height: 1;
        z-index: 3;
    }
    .pas-truck-cab {
        position: absolute;
        left: 30px;
        top: 7px;
        width: 19px;
        height: 18px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-radius: 3px 5px 3px 2px;
        z-index: 2;
    }
    .pas-truck-window {
        position: absolute;
        left: 34px;
        top: 10px;
        width: 7px;
        height: 7px;
        background: #a8d8e8;
        border: 2px solid #0A0A0A;
        border-radius: 2px;
        z-index: 4;
    }
    .pas-truck-nose {
        position: absolute;
        left: 47px;
        top: 17px;
        width: 8px;
        height: 8px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-left: none;
        border-radius: 0 3px 3px 0;
    }
    .pas-wheel {
        position: absolute;
        bottom: 0;
        width: 9px;
        height: 9px;
        background: #0A0A0A;
        border: 2px solid #222;
        border-radius: 50%;
        animation: pas-wheel-spin .32s linear infinite;
        z-index: 5;
    }
    .pas-wheel::after {
        content: "";
        position: absolute;
        inset: 2px;
        background: #FFD400;
        border-radius: 50%;
    }
    .pas-wheel.back { left: 13px; }
    .pas-wheel.front { left: 41px; }
    @keyframes pas-wheel-spin { to { transform: rotate(360deg); } }

    .pas-speed-lines { position: absolute; left: -30px; top: 17px; width: 24px; height: 18px; }
    .pas-speed-lines span { display:block; height:2px; background:#b9b9b9; margin:4px 0; border-radius:2px; animation: pas-flicker .55s linear infinite; }
    .pas-speed-lines span:nth-child(2) { width: 16px; margin-left: 8px; }
    .pas-speed-lines span:nth-child(3) { width: 11px; margin-left: 13px; }
    @keyframes pas-flicker { 50% { opacity:.25; transform: translateX(-5px); } }

    .pas-dust { position:absolute; left:-5px; bottom:0; width:34px; height:14px; opacity:.75; }
    .pas-dust span { position:absolute; bottom:0; background:#dac6a9; border-radius:50%; animation: pas-dust 1s linear infinite; }
    .pas-dust span:nth-child(1) { width:12px; height:6px; left:0; }
    .pas-dust span:nth-child(2) { width:16px; height:7px; left:10px; animation-delay:.2s; }
    .pas-dust span:nth-child(3) { width:11px; height:5px; left:23px; animation-delay:.4s; }
    @keyframes pas-dust { 50% { transform: translateX(-8px) scale(1.15); opacity:.4; } }

    .pas-stickman {
        position: absolute;
        left: 92px;
        bottom: 5px;
        width: 28px;
        height: 34px;
        animation: pas-runner-bob .35s ease-in-out infinite alternate;
    }
    @keyframes pas-runner-bob { from { transform: translateY(1px); } to { transform: translateY(-2px); } }
    .pas-stick-head {
        position:absolute;
        top:0;
        left:11px;
        width:8px;
        height:8px;
        border:2px solid #111;
        border-radius:50%;
        background:white;
    }
    .pas-stick-body { position:absolute; left:15px; top:9px; width:2px; height:13px; background:#111; transform: rotate(12deg); transform-origin:top; }
    .pas-stick-arm-a, .pas-stick-arm-b, .pas-stick-leg-a, .pas-stick-leg-b { position:absolute; width:2px; height:12px; background:#111; transform-origin:top; border-radius:2px; }
    .pas-stick-arm-a { left:15px; top:11px; transform: rotate(58deg); animation: pas-arm-a .35s linear infinite alternate; }
    .pas-stick-arm-b { left:15px; top:11px; transform: rotate(-50deg); animation: pas-arm-b .35s linear infinite alternate; }
    .pas-stick-leg-a { left:16px; top:21px; height:14px; transform: rotate(48deg); animation: pas-leg-a .35s linear infinite alternate; }
    .pas-stick-leg-b { left:16px; top:21px; height:14px; transform: rotate(-42deg); animation: pas-leg-b .35s linear infinite alternate; }
    @keyframes pas-arm-a { to { transform: rotate(-45deg); } }
    @keyframes pas-arm-b { to { transform: rotate(55deg); } }
    @keyframes pas-leg-a { to { transform: rotate(-45deg); } }
    @keyframes pas-leg-b { to { transform: rotate(48deg); } }
    </style>
    """,
    unsafe_allow_html=True,
)


st.markdown(
    f"""
    <style>
    /* ===== PAS V2 target layout overrides: safe Streamlit-native controls ===== */
    .stApp {{ background: #f7f8fa !important; color: #0A0A0A !important; font-family: Inter, "Segoe UI", Arial, sans-serif; }}
    .block-container {{ max-width: 1580px !important; padding-top: 1.05rem !important; padding-left: 2rem !important; padding-right: 2rem !important; padding-bottom: 2rem !important; }}

    section[data-testid="stSidebar"] {{ background: linear-gradient(180deg, #050606 0%, #0b1015 100%) !important; border-right: 1px solid #161b22; }}
    section[data-testid="stSidebar"] > div:first-child {{ padding-top: 1.05rem !important; }}
    section[data-testid="stSidebar"] img {{ border-radius: 14px !important; box-shadow: 0 10px 24px rgba(0,0,0,.26); }}
    .pas-sidebar-title {{ color:#fff; font-size:18px; font-weight:950; line-height:1.15; text-align:center; margin: 20px 0 8px; }}
    .pas-yellow-line {{ width:72px; height:4px; background:{PAS_YELLOW}; border-radius:99px; margin: 0 auto 22px; }}
    .pas-sidebar-copy {{ color:#fff !important; font-size:14px; line-height:1.52; font-weight:650; margin-bottom:24px; }}
    .pas-sidebar-rule {{ border-top:1px solid rgba(255,255,255,.22); margin:22px 0; }}
    .pas-sidebar-heading {{ color:{PAS_YELLOW}; font-size:19px; font-weight:950; margin: 0 0 16px; }}
    .pas-nav-row {{ display:grid; grid-template-columns: 26px 1fr; gap:10px; align-items:start; margin: 15px 0; color:#fff; font-weight:750; line-height:1.25; font-size:14px; }}
    .pas-nav-icon svg {{ width:21px; height:21px; stroke:{PAS_YELLOW}; stroke-width:2.4; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .pas-sidebar-footer {{ color:#fff; font-size:12px; font-weight:800; margin-top:28px; }}

    .pas-hero {{ display:flex; align-items:center; gap:16px; background: linear-gradient(100deg, #08090b 0%, #151718 70%, #c9aa00 130%) !important; border-radius: 16px !important; padding: 12px 22px !important; margin: 0 0 18px 0 !important; box-shadow: 0 9px 25px rgba(0,0,0,.13) !important; min-height:60px; }}
    .pas-hero-logo {{ width:37px; height:37px; border-radius:7px; background:{PAS_YELLOW}; color:#000; display:inline-flex; align-items:center; justify-content:center; font-weight:950; font-size:14px; letter-spacing:-1px; }}
    .pas-hero-text {{ color:#fff; font-size:18px; font-weight:950; letter-spacing:-.02em; }}
    .pas-hero-dot {{ color:#fff; opacity:.8; margin: 0 7px; }}
    .pas-hero-version {{ color:{PAS_YELLOW}; font-weight:950; }}

    .pas-upload-card {{ background:#fff; border:1px solid #e5e7eb; border-radius:18px; box-shadow:0 5px 18px rgba(15,23,42,.08); padding:16px 18px 14px; margin-bottom:14px; }}
    .pas-upload-title {{ color:#0A0A0A; font-size:16px; font-weight:950; margin-bottom:10px; }}
    div[data-testid="stFileUploader"] {{ margin:0 !important; }}
    div[data-testid="stFileUploader"] label {{ display:none !important; }}
    div[data-testid="stFileUploader"] section {{ background:#f4f6f8 !important; border:1px solid #dfe4ea !important; border-radius:11px !important; min-height:52px !important; padding:8px 10px !important; }}
    div[data-testid="stFileUploader"] section * {{ color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] button {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #d7dce3 !important; border-radius:10px !important; font-weight:900 !important; box-shadow:0 2px 8px rgba(0,0,0,.06) !important; }}
    div[data-testid="stFileUploader"] svg {{ color:#0A0A0A !important; fill:currentColor !important; stroke:currentColor !important; }}
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ background:#fff !important; border:1px solid #dfe4ea !important; border-radius:10px !important; color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] small {{ color:#4b5563 !important; }}

    div.stButton > button[kind="secondary"], .stButton > button {{ min-height:52px !important; font-size:16px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}
    .stDownloadButton > button {{ min-height:62px !important; font-size:20px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}

    .kpi-card {{ background:#fff !important; border-radius:18px !important; border:1px solid #e4e7eb !important; box-shadow:0 5px 20px rgba(15,23,42,.08) !important; min-height:118px !important; padding:18px 22px !important; display:flex; align-items:center; gap:18px; }}
    .kpi-icon {{ width:64px; height:64px; border-radius:50%; background:#fff5bd; display:flex; align-items:center; justify-content:center; flex:none; }}
    .kpi-icon svg {{ width:35px; height:35px; stroke:#0A0A0A; stroke-width:2.5; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .kpi-label {{ color:#111 !important; font-size:15px !important; font-weight:950 !important; margin:0 0 3px !important; }}
    .kpi-value {{ color:#e9b900 !important; font-size:42px !important; line-height:.98 !important; font-weight:950 !important; text-shadow:none !important; }}
    .kpi-sub {{ color:#374151 !important; font-size:14px !important; margin-top:6px !important; }}
    .kpi-unmatched .kpi-value {{ color:#e12626 !important; }}
    .kpi-matched .kpi-value {{ color:#16a34a !important; }}

    .pas-results-title {{ color:#0A0A0A !important; font-size:28px !important; font-weight:950 !important; margin: 22px 0 8px !important; }}
    .pas-unmatched-pill {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:0 !important; border-radius:14px 14px 0 0 !important; padding:13px 20px !important; font-size:18px; box-shadow:0 4px 14px rgba(0,0,0,.09); }}
    .pas-table-wrap {{ background:#fff !important; border:1px solid #e0e4e9 !important; border-radius:0 16px 16px 16px !important; max-height:430px !important; overflow:auto !important; box-shadow:0 7px 25px rgba(15,23,42,.10) !important; }}
    table.pas-table {{ font-size:14px !important; color:#0A0A0A !important; }}
    table.pas-table thead th {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:1px solid #e2ba00 !important; padding:12px 14px !important; font-weight:950 !important; position:sticky; top:0; z-index:5; }}
    table.pas-table tbody td {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #e1e5eb !important; padding:10px 14px !important; }}
    table.pas-table tbody tr:nth-child(even) td {{ background:#fbfcfd !important; }}
    .pas-pdf-icon {{ display:inline-flex; align-items:center; justify-content:center; width:17px; height:20px; background:#e11d2e; color:#fff; border-radius:3px; font-size:9px; font-weight:950; margin-right:8px; vertical-align:middle; }}
    table.pas-table a {{ color:#006bd6 !important; font-weight:850 !important; }}
    table.pas-table .query-cell {{ min-width:120px; white-space:nowrap; }}
    .pas-note, .pas-support, .pas-support * {{ color:#0A0A0A !important; }}
    .pas-support {{ margin-top:22px !important; font-size:15px !important; }}
    .pas-support a {{ color:#006bd6 !important; font-weight:900 !important; margin-left:12px; }}

    /* --- uploader chip cleanup: hide Streamlit's ugly uploaded-file pill and use our own card --- */
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ display: none !important; }}
    div[data-testid="stFileUploaderDropzone"] {{ background: transparent !important; border: 0 !important; padding: 0 !important; min-height: 0 !important; }}
    div[data-testid="stFileUploaderDropzoneInstructions"] {{ display: none !important; }}
    div[data-testid="stFileUploader"] section {{ background: transparent !important; border: 0 !important; min-height: 0 !important; padding: 0 !important; }}
    div[data-testid="stFileUploader"] button {{
        background: #ffffff !important;
        color: #0A0A0A !important;
        border: 1px solid #d7dce3 !important;
        border-radius: 10px !important;
        font-weight: 900 !important;
        min-height: 44px !important;
        box-shadow: 0 2px 8px rgba(0,0,0,.06) !important;
    }}
    .pas-file-card {{
        display:flex; align-items:center; gap:14px;
        background:#f4f6f8; border:1px solid #dfe4ea; border-radius:12px;
        padding:11px 14px; min-height:54px; margin: 4px 0 12px;
    }}
    .pas-file-icon {{ width:32px; height:32px; border-radius:8px; display:flex; align-items:center; justify-content:center; color:#fff; font-weight:950; font-size:11px; box-shadow:0 2px 8px rgba(0,0,0,.12); flex:none; }}
    .pas-file-icon.excel {{ background:#118a3b; }}
    .pas-file-icon.pdf {{ background:#df1f2d; }}
    .pas-file-main {{ flex:1; min-width:0; }}
    .pas-file-name {{ color:#0A0A0A; font-weight:950; font-size:15px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .pas-file-size {{ color:#4b5563; font-weight:650; font-size:13px; margin-top:2px; }}
    .pas-file-check {{ width:24px; height:24px; border-radius:50%; background:#108a37; color:white; display:flex; align-items:center; justify-content:center; font-size:15px; font-weight:950; flex:none; }}
    </style>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.image("pas_logo.png", use_column_width=True)
    st.markdown(
        """
        <div class="pas-sidebar-title">PAS Plant<br>Invoice Matching</div>
        <div class="pas-yellow-line"></div>
        <div class="pas-sidebar-copy">Upload the Plant workbook and invoice PDFs/ZIP, then export a clean reconciliation workbook.</div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-heading">Instructions</div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M16 16l-4-4-4 4"/><path d="M12 12v9"/><path d="M20 16.6A5 5 0 0 0 18 7h-1.3A8 8 0 1 0 4 15.3"/></svg></span><span>Upload Hire Order<br>Spreadsheet</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><path d="M14 2v6h6"/><path d="M9 13h6"/><path d="M9 17h6"/></svg></span><span>Upload ZIP of all invoices<br>to be checked</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M5 3l14 9-14 9V3z"/></svg></span><span>Run Reconciliation</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><path d="M7 10l5 5 5-5"/><path d="M12 15V3"/></svg></span><span>Download Reconciliation<br>Spreadsheet</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.3-4.3"/></svg></span><span>Smoke Crack</span></div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-footer">PAS NW Ltd • v1.0 Prototype Build</div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <div class="pas-hero">
      <div class="pas-hero-logo">PAS</div>
      <div class="pas-hero-text">PAS NW Ltd<span class="pas-hero-dot">•</span><span class="pas-hero-version">v1.0 Prototype Build</span></div>
    </div>
    """,
    unsafe_allow_html=True,
)



def render_bottom_chase():
    """Small non-intrusive PAS dump truck chase animation pinned to the bottom of the white content area."""
    st.markdown(
        """
        <div class="pas-bottom-chase-wrap" aria-hidden="true">
            <div class="pas-bottom-ground"></div>
            <div class="pas-chase-pack">
                <div class="pas-speed-lines"><span></span><span></span><span></span></div>
                <div class="pas-dust"><span></span><span></span><span></span></div>
                <div class="pas-truck-mini">
                    <div class="pas-truck-bed"></div>
                    <div class="pas-truck-logo">PAS</div>
                    <div class="pas-truck-cab"></div>
                    <div class="pas-truck-window"></div>
                    <div class="pas-truck-nose"></div>
                    <div class="pas-wheel back"></div>
                    <div class="pas-wheel front"></div>
                </div>
                <div class="pas-stickman">
                    <div class="pas-stick-head"></div>
                    <div class="pas-stick-body"></div>
                    <div class="pas-stick-arm-a"></div>
                    <div class="pas-stick-arm-b"></div>
                    <div class="pas-stick-leg-a"></div>
                    <div class="pas-stick-leg-b"></div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

BAD_SUPPLIER_PATTERNS = [
    "invoice", "invoice no", "invoice number", "vat", "vat number", "account", "account no",
    "customer", "customer ref", "your ref", "order no", "page", "date", "due date",
    "delivery address", "site address", "ship to", "bill to", "hire period", "total charge",
    "weekly rate", "payment", "bank", "sort code", "goods total", "invoice total", "net total",
    "pas (nw)", "pas nw", "p a s", "pocket nook", "lowton", "warrington",
]

MOVEMENT_TERMS = ["delivery", "collection", "haulage", "transport", "low loader", "lowloader", "uplift", "cartage", "movement", "carriage"]
ADDRESS_TERMS = ["delivery address", "delivery site", "site address", "ship to", "delivery details", "delivery contact", "collection address"]

STATUS_WORDS = ["on hire", "off-hired", "off hired", "purchase", "repair", "maintenance", "damage", "charges", "movement", "operated plant"]


def clean_cell(value) -> str:
    if pd.isna(value):
        return ""
    value = str(value).strip()
    if value.lower() in ["nan", "none", "nat"]:
        return ""
    return value


def norm(value) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean_cell(value).upper())


def money_to_float(value) -> Optional[float]:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value)
    match = re.search(r"-?£?\s*([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)(?:\.([0-9]{2}))?", text)
    if not match:
        return None
    whole = match.group(1).replace(",", "")
    dec = match.group(2) or "00"
    try:
        return float(f"{whole}.{dec}")
    except Exception:
        return None


def find_col(columns: List[str], keywords: List[str]) -> Optional[str]:
    norm_cols = {c: re.sub(r"[^a-z0-9]+", "", str(c).lower()) for c in columns}
    for key in keywords:
        nkey = re.sub(r"[^a-z0-9]+", "", key.lower())
        for col, ncol in norm_cols.items():
            if nkey == ncol:
                return col
    for key in keywords:
        nkey = re.sub(r"[^a-z0-9]+", "", key.lower())
        for col, ncol in norm_cols.items():
            if nkey in ncol:
                return col
    return None


def build_order_ref(job: str, order: str) -> str:
    job = clean_cell(job).upper().replace(" ", "")
    order = clean_cell(order).upper().replace(" ", "")
    if not job and not order:
        return ""
    if "/" in order and not job:
        return order
    if "/" in job and not order:
        return job
    if job and order:
        if order.startswith(job):
            return order
        return f"{job}/{order}"
    return job or order


def normalize_order_ref(ref: str) -> str:
    """Normalise PAS order refs for phased job codes.

    Operational rule:
    - P153G1/H7697 and P153G2/H7697 should match P153/H7697.
    - P151M&N/H7516 or P151MN/H7516 should match P151/H7516.

    The H number remains exact; only the job-code phase suffix is stripped.
    """
    ref = clean_cell(ref).upper().replace(" ", "")
    ref = ref.replace("\\", "/")
    if not ref:
        return ""
    # Convert P153G2H7697 to P153G2/H7697 if the slash is missing.
    ref = re.sub(r"^(P\d{3,4}[A-Z0-9&]*)(H\d{3,6})$", r"\1/\2", ref)
    m = re.match(r"^(P\d{3,4})([A-Z0-9&]*)/(H\d{3,6})$", ref)
    if not m:
        return norm(ref)
    base, suffix, hire_no = m.groups()
    # Strip known phase/group suffixes. Keep this deliberately conservative.
    suffix_clean = suffix.replace("&", "")
    if suffix_clean in {"G1", "G2", "MN"}:
        return norm(f"{base}/{hire_no}")
    return norm(f"{base}{suffix}/{hire_no}")


def load_plant_workbook(uploaded_file) -> Tuple[pd.DataFrame, Dict[str, str]]:
    xls = pd.ExcelFile(uploaded_file)
    sheet = "Plant" if "Plant" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sheet)
    df = df.dropna(how="all").copy()
    columns = list(df.columns)
    colmap = {
        "supplier": find_col(columns, ["Supplier", "Vendor", "Hire Supplier"]),
        "description": find_col(columns, ["Description", "Item Description", "Plant Description"]),
        "fleet": find_col(columns, ["Fleet No", "Fleet", "Asset No", "Item No"]),
        "cost": find_col(columns, ["Cost", "Rate", "Weekly Rate", "Price", "Value"]),
        "delivery": find_col(columns, ["Delivery", "Delivery Cost", "Haulage", "Transport"]),
        "status": find_col(columns, ["Status"]),
        "job": find_col(columns, ["Job No", "Job", "Project No"]),
        "order": find_col(columns, ["Order Number", "Order No", "Hire No", "PO", "PO Number"]),
        "on_hire": find_col(columns, ["On Hire / Delivery Date", "On Hire Date", "Delivery Date"]),
        "off_hire": find_col(columns, ["Off Hire Date", "Actual Off Hire Date"]),
        "expected_off_hire": find_col(columns, ["Expected Off Hire Date"]),
    }

    df["__row_number"] = df.index + 2
    df["__supplier"] = df[colmap["supplier"]].apply(clean_cell) if colmap["supplier"] else ""
    df["__description"] = df[colmap["description"]].apply(clean_cell) if colmap["description"] else ""
    df["__fleet"] = df[colmap["fleet"]].apply(clean_cell) if colmap["fleet"] else ""
    df["__status"] = df[colmap["status"]].apply(clean_cell) if colmap["status"] else ""
    df["__cost"] = df[colmap["cost"]].apply(money_to_float) if colmap["cost"] else None
    df["__delivery"] = df[colmap["delivery"]].apply(money_to_float) if colmap["delivery"] else None
    if colmap["job"] or colmap["order"]:
        df["__order_ref"] = df.apply(lambda r: build_order_ref(r.get(colmap["job"], ""), r.get(colmap["order"], "")), axis=1)
    else:
        df["__order_ref"] = ""
    df["__norm_order"] = df["__order_ref"].apply(normalize_order_ref)
    return df, colmap


def extract_text_from_pdf(file_bytes: bytes) -> List[str]:
    if PdfReader is None:
        raise RuntimeError("PDF reader is not available. Add pypdf or PyPDF2 to requirements.txt")
    reader = PdfReader(io.BytesIO(file_bytes))
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            pages.append("")
    return pages


def extract_invoice_number(text: str) -> str:
    """Extract invoice number without accidentally swallowing header labels.

    The PDF text order is messy on some suppliers. Boundary, for example, returns:
    date -> invoice number -> Hire Invoice -> labels. We therefore use strict
    candidate validation and supplier-neutral fallback patterns.
    """
    text = text.replace("\xa0", " ")

    # SLD / Carrier invoices often show the real number as *HI312614* while
    # "Hire Invoice No" is visually far away from the number in extracted text.
    m_hi = re.search(r"\*\s*HI\s*(\d{5,8})\s*\*", text, flags=re.IGNORECASE)
    if m_hi:
        return m_hi.group(1)

    # Filename/text style with a plain hire invoice number next to the HI marker.
    m_hi2 = re.search(r"\bHI\s*(\d{5,8})\b", text, flags=re.IGNORECASE)
    if m_hi2:
        return m_hi2.group(1)

    def valid(candidate: str) -> Optional[str]:
        candidate = re.sub(r"[^A-Z0-9\-/]", "", str(candidate).upper())
        if len(candidate) < 4 or len(candidate) > 24:
            return None
        bad_bits = ["INVOICE", "DATE", "ACCOUNT", "CUSTOMER", "PERIOD", "CHARGE", "TOTAL", "NUMBER", "CONTRACT", "ADDRESS", "SLDBOLTON", "PASNW", "PASNWLTD", "POCKET", "NOOK", "LANE", "ROAD", "STREET", "AVENUE", "WARRINGTON", "LOWTON", "ADDRESS"]
        if any(b in candidate for b in bad_bits):
            return None
        if candidate in ["UNKNOWN", "HIRE", "NO", "N/A"]:
            return None
        return candidate

    patterns = [
        r"INVOICE\s+NO\s*[:\-]\s*(\d{4,10})",
        r"Invoice\s*(?:No\.?|Number|#)\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-/]{3,24})",
        r"INVOICE\s*NO\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-/]{3,24})",
        r"Hire\s+Invoice\s+No\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-/]{3,24})",
        r"#\s*([A-Z]{2,}[A-Z0-9\-/]{4,24})",
    ]
    for p in patterns:
        m = re.search(p, text, flags=re.IGNORECASE)
        if m:
            cand = valid(m.group(1))
            if cand:
                return cand

    # Boundary-style PDFs often show a standalone 5/6 digit invoice number before "Hire Invoice".
    m = re.search(r"(?m)^\s*(\d{5,8})\s*$\s*\n\s*Hire\s+Invoice", text, flags=re.IGNORECASE)
    if m:
        return m.group(1)

    # If labels are on separate lines, scan a few lines after an invoice label for a standalone number/code.
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for i, line in enumerate(lines[:60]):
        if re.search(r"invoice\s*(no|number|#)", line, re.I):
            for nxt in lines[i + 1 : i + 8]:
                cand = valid(nxt)
                if cand:
                    return cand

    # Final supplier-neutral fallback: a standalone numeric invoice-like line near top of page.
    for line in lines[:25]:
        m = re.fullmatch(r"\d{5,8}", line)
        if m:
            return line
    return "Unknown"


def bad_invoice_number(candidate: str) -> bool:
    candidate = clean_cell(candidate).upper()
    if not candidate or candidate == "UNKNOWN":
        return True
    bad_bits = [
        "INVOICE", "DATE", "ACCOUNT", "CUSTOMER", "PERIOD", "CHARGE", "TOTAL",
        "NUMBER", "CONTRACT", "ADDRESS", "POCKET", "NOOK", "LANE", "ROAD",
        "STREET", "AVENUE", "WARRINGTON", "LOWTON", "PAS", "SLDBOLTON",
        "WA31AB", "ORDER", "GREATER", "AINSCOUGH", "HAULAGE", "CHEMI",
        "ACCOUNTNO", "INVOICEDATE", "ACCOUNTNOP105"
    ]
    return any(bit in candidate for bit in bad_bits)

def invoice_number_from_filename(filename: str) -> str:
    """Return a strong invoice number from the source PDF filename.

    In practice, most single-invoice PDFs are named after the invoice number.
    This is more reliable than PDF text extraction for suppliers where labels
    are read in the wrong order, e.g. Kensite/Boundary headers.
    Multi-invoice PDFs usually have descriptive names and will not match these
    patterns, so their invoice numbers still come from page text.
    """
    raw = str(filename or "").split(" / record")[0]
    base = Path(raw).stem.upper()
    base = re.sub(r"[_]+", " ", base)

    # Strong alphanumeric invoice formats first.
    patterns = [
        r"\b(I\d{3}I\d{5,})\b",      # GSF, e.g. I261I022140
        r"\b(WIN\d{4,8})\b",          # Ashley
        r"\b(INV\d{4,9})\b",          # Fox etc
        r"\b(\d{5,8})Q\b",            # Kensite filenames like 915708Q ON SAGE
        r"\b(\d{5,8})\b",             # Boundary/RSS/SLD/SMT/Tyrefix
    ]
    for pat in patterns:
        m = re.search(pat, base)
        if m:
            return m.group(1)
    return ""


def choose_invoice_number(filename: str, extracted: str) -> str:
    file_no = invoice_number_from_filename(filename)
    extracted = clean_cell(extracted) or "Unknown"
    extracted_upper = extracted.upper()

    # If the filename contains a strong invoice number, prefer it. This avoids
    # glued header text such as INVOICEDATEACCOUNTNO and branch/customer lines.
    if file_no:
        if (
            extracted_upper == "UNKNOWN"
            or bad_invoice_number(extracted)
            or len(extracted_upper) > 18
            or not re.search(r"\d", extracted_upper)
            or extracted_upper in {"INVOICE", "HIRE INVOICE", "SALES INVOICE"}
        ):
            return file_no
        # If extracted is merely the filename number with extra junk attached, use filename.
        if extracted_upper.startswith(file_no) and extracted_upper != file_no:
            return file_no
        # For standard single-invoice files, filename is normally the safest source.
        if re.fullmatch(r"(\d{5,8}|WIN\d{4,8}|INV\d{4,9}|I\d{3}I\d{5,})", file_no):
            return file_no

    return extracted


def extract_invoice_date(text: str) -> str:
    """Extract the document invoice date, not hire/delivery/due/payment dates.

    Priority is given only to dates that sit next to strong document-date labels.
    This prevents hire-period dates such as 18/05/2026 - 24/05/2026 or due dates
    being pulled through as the invoice date.
    """
    def norm_date(raw: str) -> str:
        return clean_cell(raw).replace("  ", " ")

    date_token = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}[-\s]?[A-Za-z]{3,9}[-\s]?\d{2,4})"

    compact = re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()
    compact_patterns = [
        rf"Date\s+of\s+Invoice\s*[:\-]?\s*{date_token}",
        rf"Invoice\s+Date\s*[:\-]?\s*{date_token}",
        rf"Document\s+Date\s*[:\-]?\s*{date_token}",
        rf"Date\s*[:\-]\s*{date_token}",
    ]
    for pat in compact_patterns:
        m = re.search(pat, compact, re.I)
        if m:
            return norm_date(m.group(1))

    # Boundary text often extracts as: '24 May 2026 176607 Hire Invoice'.
    m = re.search(rf"{date_token}\s+\d{{5,8}}\s+Hire\s+Invoice", compact, re.I)
    if m:
        return norm_date(m.group(1))

    strong_patterns = [
        rf"Date\s+of\s+Invoice\s*[:\-]?\s*{date_token}",
        rf"Invoice\s+Date\s*[:\-]?\s*{date_token}",
        rf"Invoice\s+dated\s*[:\-]?\s*{date_token}",
        rf"Document\s+Date\s*[:\-]?\s*{date_token}",
    ]
    for pat in strong_patterns:
        m = re.search(pat, text, re.I)
        if m:
            return norm_date(m.group(1))

    # Some suppliers put the invoice date on the line immediately after the label.
    lines = [clean_cell(l) for l in text.splitlines() if clean_cell(l)]
    for i, line in enumerate(lines[:80]):
        if re.search(r"^(Date\s+of\s+Invoice|Invoice\s+Date|Document\s+Date)\b", line, re.I):
            window = " ".join(lines[i:i+4])
            m = re.search(date_token, window, re.I)
            if m:
                return norm_date(m.group(1))

    # Weak fallback: only accept a generic 'Date:' near the invoice header, and reject known non-document labels.
    for i, line in enumerate(lines[:60]):
        if re.search(r"\b(Due|Hire|Delivery|Collection|From|To|Payment|Period)\b", line, re.I):
            continue
        m = re.search(rf"^Date\s*[:\-]?\s*{date_token}$", line, re.I)
        if m:
            return norm_date(m.group(1))

    return ""


def split_pdf_into_invoices(filename: str, pages: List[str]) -> List[Dict]:
    groups = []
    current = None
    for idx, page_text in enumerate(pages, start=1):
        raw_inv_no = extract_invoice_number(page_text)
        inv_no = choose_invoice_number(filename, raw_inv_no)
        has_invoice_header = bool(re.search(r"invoice\s*(no\.?|number|#)", page_text, re.I))
        page_marker = re.search(r"Page\s+\d+\s*(?:/|of)\s*\d+", page_text, re.I)

        if current is None:
            current = {"source_file": filename, "invoice_number": inv_no, "pages": [idx], "text": page_text}
            continue

        # Start a new invoice only where a genuine new invoice number is found, not just continuation page text.
        current_inv = current.get("invoice_number", "Unknown")
        if has_invoice_header and inv_no != "Unknown" and inv_no != current_inv:
            groups.append(current)
            current = {"source_file": filename, "invoice_number": inv_no, "pages": [idx], "text": page_text}
        else:
            current["pages"].append(idx)
            current["text"] += "\n" + page_text
            if current_inv == "Unknown" and inv_no != "Unknown":
                current["invoice_number"] = inv_no

    if current:
        groups.append(current)
    return groups


def extract_order_refs(text: str) -> List[str]:
    refs = set()
    # Full PAS order references.
    for m in re.finditer(r"\b(P\d{3,4}[A-Z0-9&]*\s*/\s*H\d{3,6})\b", text, re.I):
        refs.add(m.group(1).upper().replace(" ", ""))
    # Some suppliers remove slash or ampersand spacing e.g. P151MN/H7516.
    for m in re.finditer(r"\b(P\d{3,4}[A-Z0-9&]*\s*H\d{3,6})\b", text, re.I):
        raw = m.group(1).upper().replace(" ", "")
        if "/" not in raw:
            raw = re.sub(r"(P\d{3,4}[A-Z&]*)(H\d+)", r"\1/\2", raw)
        refs.add(raw)
    return sorted(refs)


def extract_account_name_supplier(text: str) -> str:
    """
    Some invoices, especially Boundary Plant, show the clean supplier name in the
    payment details block as 'Account Name: Boundary Plant Hire Ltd'. Use this as
    a high-confidence supplier source and ignore customer/branch/address labels.
    """
    compact = re.sub(r"\s+", " ", text).strip()
    m = re.search(
        r"Account\s+Name\s*:?\s*([A-Z0-9&.,'()\-/ ]{3,90}?)(?=\s+(Sort\s+Code|Account\s+No|IBAN|Bank|Payment|VAT|$))",
        compact,
        re.I,
    )
    if not m:
        return ""
    supplier = re.sub(r"\s+", " ", m.group(1)).strip(" :-")
    if any(bad in supplier.lower() for bad in BAD_SUPPLIER_PATTERNS):
        return ""
    if len(supplier) < 4:
        return ""
    return supplier


def extract_known_supplier_from_text(text: str) -> str:
    """High-confidence supplier extraction from explicit supplier identifiers.

    This is not matching logic. It only stops the app from using customer/address/table
    text as a supplier where the actual supplier name is clearly printed somewhere on
    the invoice PDF.
    """
    compact = re.sub(r"\s+", " ", text).strip()
    low = compact.lower()

    # Payment/account name blocks are usually the cleanest source.
    account_supplier = extract_account_name_supplier(text)
    if account_supplier:
        if "boundary plant" in account_supplier.lower():
            return "Boundary Plant"
        return account_supplier

    # Common printed supplier names/logos/footer identifiers.
    if "kensite" in low or "kensite services" in low:
        return "Kensite"
    if "sld bolton" in low or "sldpumpspower" in low or "carrier rental systems" in low:
        return "SLD Pumps"
    if "smiths equipment hire" in low or "smithshire" in low:
        return "Smiths Hire"
    if "ashley plant hire" in low:
        return "Ashley Plant"
    if "rope & sling specialists" in low or "rssgroup" in low:
        return "Rope & Sling Specialists Ltd"
    if "gsf car parts" in low:
        return "GSF Car Parts"
    if "tyrefix plant tyres" in low:
        return "Tyrefix Plant Tyres (UK) Ltd"
    if "fox brothers" in low or "fox group" in low:
        return "Fox Brothers"
    if "smt gb" in low or "smt.network" in low or "services machinery" in low:
        return "SMT GB"
    return ""


def extract_supplier_candidate(text: str) -> str:
    # Use high-confidence supplier identifiers before any generic header guess.
    known_supplier = extract_known_supplier_from_text(text)
    if known_supplier:
        return known_supplier

    lines = [re.sub(r"\s+", " ", l).strip() for l in text.splitlines() if l.strip()]
    candidates = []
    for line in lines[:35]:
        low = line.lower()
        if any(bad in low for bad in BAD_SUPPLIER_PATTERNS):
            continue
        if re.fullmatch(r"[0-9\- ]+", line):
            continue
        if len(line) < 4 or len(line) > 80:
            continue
        # Strong company-style lines.
        if re.search(r"\b(ltd|limited|plc|llp|group|hire|plant|parts|equipment|specialists|services|systems|reclamation)\b", line, re.I):
            candidates.append(line)
    if candidates:
        return candidates[0]
    # Fallback: first sensible top-left/header line.
    for line in lines[:12]:
        low = line.lower()
        if not any(bad in low for bad in BAD_SUPPLIER_PATTERNS) and len(line) >= 4:
            return line
    return "Unknown"


def extract_money_values(text: str) -> List[float]:
    values = []
    for line in text.splitlines():
        low = line.lower()
        # Prefer line-level/rate values. Still include net total as low priority later if no line values match.
        matches = re.findall(r"£\s*([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)(?:\.([0-9]{2}))?", line)
        if not matches:
            # Boundary-style rows sometimes include 200.00 without £.
            if re.search(r"\b(rate|weekly|week|wk|total charge|unit price|net|value|amount|price)\b", low):
                matches = re.findall(r"\b([0-9]{1,4})(?:\.([0-9]{2}))\b", line)
        for whole, dec in matches:
            try:
                values.append(float(f"{whole.replace(',', '')}.{dec or '00'}"))
            except Exception:
                pass
    # De-dupe while preserving order.
    out = []
    for v in values:
        if v not in out:
            out.append(v)
    return out


def extract_priority_money_values(text: str) -> List[float]:
    """Extract comparable invoice rates/values globally, but avoid header/date noise.

    We deliberately do not require supplier-specific table layouts. If a full PO
    has matched the Plant tab, any proper monetary/rate value on the invoice can
    validate against the agreed Plant value. This fixes Boundary-style tables
    where weekly rates/total charges appear as plain decimals without a £ sign.
    """
    priority = []
    fallback = []
    for raw_line in text.splitlines():
        line = raw_line.replace("\xa0", " ")
        low = line.lower()
        # Ignore address/header labels that are not charge lines.
        if any(addr in low for addr in ADDRESS_TERMS):
            continue
        if re.search(r"\b(vat registration|company number|account no|sort code|telephone|postcode|page\s+\d)\b", low):
            continue

        line_vals = []

        # Currency values e.g. £425.00
        for m in re.finditer(r"£\s*([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)(?:\.([0-9]{2}))?", line):
            whole, dec = m.group(1), m.group(2) or "00"
            try:
                line_vals.append(float(f"{whole.replace(',', '')}.{dec}"))
            except Exception:
                pass

        # Plain decimal money/rate values e.g. Boundary rows: 425.00, 200.00.
        # Avoid VAT percentage values like 20.00%.
        for m in re.finditer(r"(?<![A-Z0-9/])([0-9]{1,5}\.[0-9]{2})(?!\s*%)(?![A-Z0-9/])", line):
            try:
                val = float(m.group(1))
                if 0 < val <= 100000:
                    line_vals.append(val)
            except Exception:
                pass

        if not line_vals:
            continue

        # De-prioritise final invoice/gross/VAT totals, but keep as fallback.
        if any(label in low for label in ["invoice total", "vat total", "gross", "total vat", "vat @", "vat at", "invoice total"]):
            fallback.extend(line_vals)
        else:
            priority.extend(line_vals)

    values = priority + fallback
    out = []
    for v in values:
        v = round(float(v), 2)
        if v not in out:
            out.append(v)
    return out


def has_actual_movement_charge(text: str) -> bool:
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for line in lines:
        low = line.lower()
        if any(addr in low for addr in ADDRESS_TERMS):
            continue
        if any(term in low for term in MOVEMENT_TERMS):
            # only actual charge line if money or an amount/rate-like number is present.
            if re.search(r"£\s*\d|\b\d+\.\d{2}\b", line):
                return True
    return False


def classify_invoice(text: str, plant_status: str = "") -> str:
    status = plant_status.lower()
    if "operated" in status:
        return "Operated Plant"
    if "movement" in status:
        return "Movement"
    if "purchase" in status:
        return "Purchase"
    if "repair" in status or "maintenance" in status:
        return "Repair/Maintenance"
    if "damage" in status or "charge" in status:
        return "Damage/Charges"

    low = text.lower()
    movement = has_actual_movement_charge(text)
    if any(w in low for w in ["puncture", "tyre repair", "tyre fix", "repair", "service", "maintenance"]):
        return "Repair/Maintenance"
    if any(w in low for w in ["loss of hired equipment", "damage", "cleaning charge", "wear charge"]):
        return "Damage/Charges"
    if movement and re.search(r"hire|weekly|week|wk|hire period", low):
        return "Hire + Movement"
    if movement:
        return "Movement"
    if re.search(r"hire|weekly|week|wk|hire period", low):
        return "Hire"
    return "Purchase"


def find_matching_plant_rows(plant_df: pd.DataFrame, refs: List[str]) -> pd.DataFrame:
    if not refs:
        return plant_df.iloc[0:0]
    norm_refs = [normalize_order_ref(r) for r in refs]
    mask = plant_df["__norm_order"].isin(norm_refs)
    return plant_df[mask].copy()


def close_money(a: float, b: float, tolerance: float = 0.02) -> bool:
    try:
        return abs(round(float(a), 2) - round(float(b), 2)) <= tolerance
    except Exception:
        return False


def extract_rate_charge_lines(text: str) -> List[Dict[str, float]]:
    """Extract structured rate/charge pairs from invoice text.

    Handles common hire invoice wording such as:
    - £18.60 per week for 1 week, 2 days ... £26.04
    - £35.00 WK 0/4/2 STD £154.00
    - Boundary-style table rows with weekly rate and total charge as plain decimals.
    """
    rows = []
    flat = re.sub(r"\s+", " ", text.replace("£ ", "£")).strip()

    def add(rate, charge=None, days=None, source=""):
        try:
            rate = round(float(str(rate).replace(",", "")), 2)
        except Exception:
            return
        if rate <= 0 or rate > 100000:
            return
        c = None
        if charge is not None:
            try:
                c = round(float(str(charge).replace(",", "")), 2)
            except Exception:
                c = None
        d = None
        if days is not None:
            try:
                d = float(days)
            except Exception:
                d = None
        if c is None and d is not None:
            c = round(rate / 5 * d, 2)
        item = {"rate": rate}
        if c is not None and c > 0:
            item["charge"] = c
        if d is not None:
            item["days"] = d
        if source:
            item["source"] = source
        if item not in rows:
            rows.append(item)

    for m in re.finditer(
        r"£\s*([0-9,]+\.\d{2})\s*per\s*week\s*for\s*"
        r"(?:(\d+)\s*week[s]?)?\s*,?\s*"
        r"(?:(\d+)\s*day[s]?)?.{0,220}?£\s*([0-9,]+\.\d{2})",
        flat, re.I
    ):
        rate = m.group(1)
        weeks = int(m.group(2) or 0)
        days = int(m.group(3) or 0)
        charge = m.group(4)
        add(rate, charge, weeks * 5 + days, "per week wording")

    for m in re.finditer(
        r"£\s*([0-9,]+\.\d{2})\s*per\s*week\s*for\s*"
        r"(?:(\d+)\s*week[s]?)?\s*,?\s*"
        r"(?:(\d+)\s*day[s]?)?",
        flat, re.I
    ):
        rate = m.group(1)
        weeks = int(m.group(2) or 0)
        days = int(m.group(3) or 0)
        if weeks or days:
            add(rate, None, weeks * 5 + days, "calculated per week wording")

    for m in re.finditer(r"£?\s*([0-9,]+\.\d{2})\s*WK\s*(\d+)\s*/\s*(\d+)\s*/\s*(\d+).{0,60}?£?\s*([0-9,]+\.\d{2})", flat, re.I):
        rate = m.group(1)
        months = int(m.group(2) or 0)
        weeks = int(m.group(3) or 0)
        days = int(m.group(4) or 0)
        charge = m.group(5)
        add(rate, charge, (months * 4 * 5) + (weeks * 5) + days, "M/W/D wording")

    for line in text.splitlines():
        low = line.lower()
        if any(addr in low for addr in ADDRESS_TERMS):
            continue
        if re.search(r"\b(invoice total|vat total|goods total|sort code|account no|payment details)\b", low):
            continue
        nums = [float(x) for x in re.findall(r"(?<![A-Z0-9/])([0-9]{1,5}\.[0-9]{2})(?!\s*%)(?![A-Z0-9/])", line)]
        wd = re.search(r"\b(\d+)\s*/\s*(\d+)\b", line)
        if len(nums) >= 2 and wd:
            weeks = int(wd.group(1) or 0)
            days = int(wd.group(2) or 0)
            chargeable_days = weeks * 5 + days

            # Most hire tables are: ... weekly_rate total_charge.
            # Some suppliers, such as Synergy, extract as:
            # qty total_charge ... weeks/days weekly_rate VAT%.
            # Detect that by checking which pairing satisfies rate / 5 * chargeable_days = charge.
            standard_rate, standard_charge = nums[-2], nums[-1]
            reversed_rate, reversed_charge = nums[-1], nums[-2]

            if chargeable_days:
                expected_standard = round(standard_rate / 5 * chargeable_days, 2)
                expected_reversed = round(reversed_rate / 5 * chargeable_days, 2)
                if close_money(expected_reversed, reversed_charge, 0.03):
                    add(reversed_rate, reversed_charge, chargeable_days, "table row reversed rate/charge")
                elif close_money(expected_standard, standard_charge, 0.03):
                    add(standard_rate, standard_charge, chargeable_days, "table row rate/charge")
                else:
                    add(standard_rate, standard_charge, chargeable_days, "table row rate/charge fallback")
            else:
                add(standard_rate, standard_charge, None, "table row rate/charge")

    return rows


def calculated_invoice_values(text: str) -> List[float]:
    vals = []
    for item in extract_rate_charge_lines(text):
        vals.append(item.get("rate"))
        if item.get("charge") is not None:
            vals.append(item.get("charge"))
    out = []
    for v in vals:
        if v is None:
            continue
        v = round(float(v), 2)
        if v > 0 and v not in out:
            out.append(v)
    return out


def values_match(plant_values: List[float], invoice_values: List[float], line_items: Optional[List[Dict[str, float]]] = None, tolerance: float = 0.02) -> Tuple[bool, str, Optional[float]]:
    """Validate invoice values against Plant values.

    Important behaviour:
    - If structured invoice lines are found, validate the invoice as a set of lines, not one isolated rate.
    - This prevents one matching rate from approving an invoice that also has an unmatched extra charge.
    - Supports pro-rata weekly hire where invoices show "£x per week for y weeks, z days".
    - Supports combined Plant rows where invoice has multiple lines but the Plant row has one combined value.
    """
    plant_values = sorted(set(round(float(v), 2) for v in plant_values if v is not None and not pd.isna(v) and float(v) > 0))
    invoice_values = sorted(set(round(float(v), 2) for v in invoice_values if v is not None and not pd.isna(v) and float(v) > 0))
    line_items = line_items or []

    line_rates = []
    line_charges = []
    validated_prorata_lines = []

    for item in line_items:
        rate = item.get("rate")
        charge = item.get("charge")
        days = item.get("days")
        if rate is not None and float(rate) > 0:
            line_rates.append(round(float(rate), 2))
        if charge is not None and float(charge) > 0:
            line_charges.append(round(float(charge), 2))
        if rate is not None and charge is not None and days is not None:
            expected = round(float(rate) / 5 * float(days), 2)
            if close_money(expected, float(charge), 0.03):
                validated_prorata_lines.append(round(float(rate), 2))

    line_rates = sorted(set(line_rates))
    line_charges = sorted(set(line_charges))

    if not plant_values:
        return False, "No agreed rate/value found on Plant tab", None
    if not invoice_values and not line_rates and not line_charges:
        return False, "No comparable rate/value found on invoice", None

    plant_sum = round(sum(plant_values), 2)
    rate_sum = round(sum(line_rates), 2) if line_rates else None
    charge_sum = round(sum(line_charges), 2) if line_charges else None
    invoice_sum = round(sum(invoice_values), 2) if invoice_values else None

    # Structured line validation comes first. This is the key invoice-level safety check.
    if line_rates or line_charges:
        # Combined Plant row case: e.g. Excavator + forks on invoice, one combined value on Plant tab.
        if rate_sum is not None and close_money(plant_sum, rate_sum, tolerance):
            return True, "Combined invoice line rates matched Plant value", 0.0
        if charge_sum is not None and close_money(plant_sum, charge_sum, tolerance):
            return True, "Combined invoice line charges matched Plant value", 0.0

        # Pro-rata weekly hire case: invoice rates match Plant rates and line charges calculate correctly.
        if validated_prorata_lines:
            unmatched_rates = []
            for rate in line_rates:
                if not any(close_money(rate, pv, tolerance) for pv in plant_values):
                    unmatched_rates.append(rate)
            if not unmatched_rates:
                return True, "Weekly rates matched and pro-rata hire charges validated", 0.0

        # Multiple Plant rows / multiple invoice lines case: every invoice line rate must be present on Plant.
        if line_rates:
            unmatched_rates = []
            for rate in line_rates:
                if not any(close_money(rate, pv, tolerance) for pv in plant_values):
                    unmatched_rates.append(rate)
            if not unmatched_rates:
                return True, "Invoice line rates matched Plant values", 0.0
            return False, "Additional or unmatched invoice charge line found", None

    # Fallback: direct value/rate matching where no structured lines were found.
    for pv in plant_values:
        for iv in invoice_values:
            if close_money(pv, iv, tolerance):
                return True, "Rate/value matched", 0.0
            if pv and abs(pv - iv) / pv <= 0.005:
                return True, "Rate/value matched within rounding tolerance", round(iv - pv, 2)

    for comp in [x for x in [rate_sum, charge_sum, invoice_sum] if x is not None and x > 0]:
        if close_money(plant_sum, comp, tolerance):
            return True, "Combined invoice values matched Plant value", 0.0
        for pv in plant_values:
            if close_money(pv, comp, tolerance):
                return True, "Combined invoice values matched agreed value", 0.0

    nearest_candidates = invoice_values + line_rates + line_charges + [x for x in [rate_sum, charge_sum, invoice_sum] if x is not None and x > 0]
    if nearest_candidates:
        nearest = min((iv - pv for pv in plant_values for iv in nearest_candidates), key=lambda x: abs(x))
        return False, f"Price discrepancy: nearest variance £{nearest:.2f}", nearest
    return False, "No comparable rate/value found on invoice", None


def extract_supplier_email(text: str) -> str:
    """Find the best supplier email address from invoice PDF text.

    Priority:
    1. Real email address printed on the invoice.
    2. Ignore PAS/internal/no-reply style addresses.
    3. Prefer accounts/credit/control/hire/sales style supplier mailboxes.
    """
    if not text:
        return ""

    # Normalise OCR spacing around @ and dots where possible.
    cleaned = text.replace(" ", " ")
    cleaned = re.sub(r"\s*@\s*", "@", cleaned)
    cleaned = re.sub(r"\s*\.\s*", ".", cleaned)

    emails = re.findall(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", cleaned, flags=re.I)
    if not emails:
        return ""

    blocked_bits = [
        "pasnw.co.uk",
        "no-reply",
        "noreply",
        "donotreply",
        "do-not-reply",
        "example.com",
    ]

    def clean_email(email: str) -> str:
        return email.strip().strip(".,;:)>").lower()

    unique = []
    for email in emails:
        email = clean_email(email)
        if not email or any(bit in email for bit in blocked_bits):
            continue
        if email not in unique:
            unique.append(email)

    if not unique:
        return ""

    priority_words = [
        "accounts", "account", "credit", "invoice", "invoices",
        "hire", "sales", "admin", "office", "orders", "queries"
    ]
    for word in priority_words:
        for email in unique:
            local = email.split("@", 1)[0]
            if word in local:
                return email

    return unique[0]

def reconcile_invoice(inv: Dict, plant_df: pd.DataFrame) -> Dict:
    text = inv["text"]
    refs = extract_order_refs(text)
    matched_rows = find_matching_plant_rows(plant_df, refs)

    raw_supplier = extract_supplier_candidate(text)
    plant_supplier = ""
    plant_status = ""
    plant_rate_values = []
    plant_rows = []

    if not matched_rows.empty:
        plant_supplier = clean_cell(matched_rows.iloc[0].get("__supplier", ""))
        plant_status = clean_cell(matched_rows.iloc[0].get("__status", ""))
        plant_rows = matched_rows["__row_number"].dropna().astype(int).astype(str).tolist()
        for _, row in matched_rows.iterrows():
            cost = row.get("__cost", None)
            delivery = row.get("__delivery", None)
            if cost is not None and not pd.isna(cost) and cost > 0:
                plant_rate_values.append(float(cost))
            # delivery is only a comparable value if an actual movement charge exists on invoice.
            if has_actual_movement_charge(text) and delivery is not None and not pd.isna(delivery) and delivery > 0:
                plant_rate_values.append(float(delivery))

    supplier = plant_supplier or raw_supplier
    invoice_type = classify_invoice(text, plant_status)
    line_items = extract_rate_charge_lines(text)
    invoice_values = extract_priority_money_values(text)
    for v in calculated_invoice_values(text):
        if v not in invoice_values:
            invoice_values.append(v)
    net_total = None
    # Find net total/goods value for display.
    for p in [r"Goods\s+Total\s*£?\s*([0-9,]+\.\d{2})", r"NET\s*£\s*([0-9,]+\.\d{2})", r"Sub-Total\s*£?\s*([0-9,]+\.\d{2})", r"Net\s+value\s*\n?\s*([0-9,]+\.\d{2})"]:
        m = re.search(p, text, re.I)
        if m:
            net_total = money_to_float(m.group(1))
            break

    order_ref = refs[0] if refs else ""
    normalised_order_ref = normalize_order_ref(order_ref) if order_ref else ""
    matched = False
    reason = ""
    variance = None

    if not refs:
        reason = "No usable order reference on invoice"
    elif matched_rows.empty:
        reason = "Order reference not found on Plant tab"
    else:
        matched, reason, variance = values_match(plant_rate_values, invoice_values, line_items)

    supplier_email = extract_supplier_email(text)
    status = "Matched" if matched else "Unmatched"
    return {
        "PDF File": inv["source_file"],
        "Invoice Number": inv.get("invoice_number", "Unknown"),
        "Supplier": supplier,
        "Supplier Email": supplier_email,
        "Invoice Date": extract_invoice_date(text),
        "Order Reference": order_ref,
        "Normalised Order Reference": normalised_order_ref,
        "Reference Quality": "Full" if refs else "Missing",
        "Invoice Type": invoice_type,
        "Plant Status": plant_status,
        "Agreed Rate / Value": ", ".join(f"£{v:.2f}" for v in sorted(set(plant_rate_values))) if plant_rate_values else "",
        "Invoice Values Found": ", ".join(f"£{v:.2f}" for v in invoice_values[:12]),
        "Pro-Rata Lines Found": ", ".join([f"£{x.get('rate', 0):.2f} -> £{x.get('charge', 0):.2f}" for x in line_items[:8] if x.get('charge') is not None]),
        "Invoice Net Total": f"£{net_total:.2f}" if net_total is not None else "",
        "Matched Plant Row(s)": ", ".join(plant_rows),
        "Match Status": status,
        "Unmatched Reason": "" if matched else reason,
        "Variance": f"£{variance:.2f}" if variance is not None else "",
        "Pages": ", ".join(map(str, inv.get("pages", []))),
    }


def collect_invoice_records(invoice_uploads) -> List[Dict]:
    records = []
    for uploaded in invoice_uploads:
        name = uploaded.name
        data = uploaded.read()
        if name.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                for member in z.namelist():
                    if member.lower().endswith(".pdf") and not member.endswith("/"):
                        pdf_bytes = z.read(member)
                        pages = extract_text_from_pdf(pdf_bytes)
                        records.extend(split_pdf_into_invoices(member.split("/")[-1], pages))
        elif name.lower().endswith(".pdf"):
            pages = extract_text_from_pdf(data)
            records.extend(split_pdf_into_invoices(name, pages))
    return records


def style_excel(writer, dfs: Dict[str, pd.DataFrame]):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="FFD400")
    black_font = Font(color="000000", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, df in dfs.items():
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"
        if df.shape[0] >= 0 and df.shape[1] > 0:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = black_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
            width = min(max(max_len + 2, 10), 42)
            ws.column_dimensions[get_column_letter(idx)].width = width


OUTPUT_COLUMNS = [
    "PDF File",
    "Invoice Number",
    "Supplier",
    "Supplier Email",
    "Order Reference",
    "Invoice Type",
    "Plant Status",
    "Agreed Rate / Value",
    "Match Status",
    "Unmatched Reason",
]


def clean_output_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    out = df.copy()
    for col in OUTPUT_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    return out[OUTPUT_COLUMNS]




def make_query_email_link(row) -> str:
    """Create a mailto link for querying an unmatched invoice.

    This opens the user's default mail client, sends to the supplier email
    found on the invoice where available, CCs invoices@pasnw.co.uk,
    and prepares a supplier query email.
    """
    invoice_no = clean_cell(row.get("Invoice Number", "Unknown")) or "Unknown"
    supplier = clean_cell(row.get("Supplier", "")) or "the supplier"
    order_ref = clean_cell(row.get("Order Reference", "")) or "Not found"
    reason = clean_cell(row.get("Unmatched Reason", "")) or "Invoice did not match the Plant order record"
    pdf_file = clean_cell(row.get("PDF File", ""))
    supplier_email = clean_cell(row.get("Supplier Email", ""))

    subject = f"Invoice Query: {invoice_no}"
    body = (
        "Hi,\n\n"
        f"We are querying invoice {invoice_no}.\n\n"
        "Our invoice matching system has flagged the following issue:\n\n"
        f"Supplier: {supplier}\n"
        f"Order reference: {order_ref}\n"
        f"Reason: {reason}\n"
    )
    if pdf_file:
        body += f"PDF file: {pdf_file}\n"
    body += (
        "\nPlease can you review and confirm whether this invoice is correct, "
        "or send a revised invoice if required.\n\n"
        "Kind regards,\n\n"
        "PAS Plant Team"
    )

    to_part = quote(supplier_email) if supplier_email else ""
    return f"mailto:{to_part}?cc=invoices@pasnw.co.uk&subject={quote(subject)}&body={quote(body)}"


def add_query_email_column(df: pd.DataFrame) -> pd.DataFrame:
    """Add a clickable Query Supplier mailto link for the app display only."""
    out = clean_output_df(df).copy()
    if out.empty:
        out["Query Supplier"] = ""
        return out
    out["Query Supplier"] = out.apply(make_query_email_link, axis=1)
    cols = list(out.columns)
    # Put action link first so it is obvious to users.
    cols = ["Query Supplier"] + [c for c in cols if c != "Query Supplier"]
    return out[cols]




def render_unmatched_table(df: pd.DataFrame):
    """Render unmatched invoices as a clean white HTML table with PAS yellow headers.

    This deliberately uses plain HTML rather than Streamlit dataframe widgets so
    Streamlit cannot render stray `None`/badge values before the table.
    """
    display_df = add_query_email_column(df)
    if display_df.empty:
        st.success("No unmatched invoices. Nice one.")
        return

    cols = [
        "Query Supplier",
        "PDF File",
        "Invoice Number",
        "Supplier",
        "Supplier Email",
        "Order Reference",
        "Invoice Type",
        "Plant Status",
        "Agreed Rate / Value",
        "Unmatched Reason",
    ]
    for col in cols:
        if col not in display_df.columns:
            display_df[col] = ""

    display_df = display_df[cols].copy()
    display_df = display_df.where(pd.notna(display_df), "")
    display_df = display_df.replace({None: "", "None": "", "nan": "", "NaN": ""})

    header_html = "".join(f"<th>{escape(str(col))}</th>" for col in cols)
    body_rows = []

    for _, row in display_df.iterrows():
        cells = []
        for col in cols:
            raw_value = row.get(col, "")
            value = "" if raw_value is None or pd.isna(raw_value) else str(raw_value).strip()
            if value.lower() in {"none", "nan", "nat"}:
                value = ""

            if col == "Query Supplier":
                if value:
                    cells.append(f'<td class="query-cell"><a href="{escape(value, quote=True)}">✉ Query Supplier</a></td>')
                else:
                    cells.append('<td class="query-cell"></td>')
            elif col == "PDF File":
                cells.append(f'<td><span class="pas-pdf-icon">PDF</span>{escape(value)}</td>')
            else:
                cells.append(f"<td>{escape(value)}</td>")
        body_rows.append("<tr>" + "".join(cells) + "</tr>")

    table_html = f"""
    <div class="pas-unmatched-pill">⚠ Unmatched Invoices</div>
    <div class="pas-table-wrap">
      <table class="pas-table">
        <thead><tr>{header_html}</tr></thead>
        <tbody>{''.join(body_rows)}</tbody>
      </table>
    </div>
    <div class="pas-note">
      Showing {min(len(display_df), 10)} of {len(display_df)} unmatched invoice(s). Query Supplier opens a pre-filled email draft to the supplier email where found, with invoices@pasnw.co.uk CC'd.
    </div>
    """
    st.markdown(table_html, unsafe_allow_html=True)




def _file_size_label(uploaded_file):
    try:
        size = uploaded_file.size
    except Exception:
        try:
            pos = uploaded_file.tell()
            uploaded_file.seek(0, 2)
            size = uploaded_file.tell()
            uploaded_file.seek(pos)
        except Exception:
            size = 0
    if size >= 1024 * 1024:
        return f"{size / (1024 * 1024):.1f} MB"
    if size >= 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size} B"


def render_selected_file_card(uploaded_file, file_kind="excel"):
    if not uploaded_file:
        return
    icon_label = "XLS" if file_kind == "excel" else "PDF"
    icon_class = "excel" if file_kind == "excel" else "pdf"
    name = escape(getattr(uploaded_file, "name", "Uploaded file"))
    size = escape(_file_size_label(uploaded_file))
    st.markdown(
        f"""
        <div class="pas-file-card">
            <div class="pas-file-icon {icon_class}">{icon_label}</div>
            <div class="pas-file-main">
                <div class="pas-file-name">{name}</div>
                <div class="pas-file-size">{size}</div>
            </div>
            <div class="pas-file-check">✓</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def make_excel(summary_df, matched_df, unmatched_df, all_df) -> bytes:
    output = io.BytesIO()
    rules_df = pd.DataFrame({
        "Rule": [
            "Invoice-level approval: a full invoice is Matched only if the order reference and rate/value validate.",
            "Full PAS order references are required. Missing or weak references are Unmatched.",
            "Supplier is taken from the Plant row once a full order reference matches.",
            "Movement is only detected from actual charge lines with values, not from Delivery Address/Site Address labels.",
            "Global value extraction checks line rate, weekly rate, daily rate, unit price, line value, total charge and net totals.",
            "Weekly hire charges are validated using pro-rata week/day calculations where the invoice shows them.",
            "Job-code phases are normalised for matching: P153G1/P153G2 become P153, and P151M&N/P151MN become P151.",
            "Raw Text Preview is excluded from the operational output.",
            "Excel exports use filters, frozen headers and auto-sized columns.",
        ]
    })
    dfs = {
        "Summary": summary_df,
        "Matched": clean_output_df(matched_df),
        "Unmatched": clean_output_df(unmatched_df),
        "All Extracted Invoices": clean_output_df(all_df),
        "Rules": rules_df,
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            df.to_excel(writer, index=False, sheet_name=sheet)
        style_excel(writer, dfs)
    return output.getvalue()


up_col1, up_col2 = st.columns(2)
with up_col1:
    st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Material Spreadsheet</div>', unsafe_allow_html=True)
    plant_file = st.file_uploader("Upload Material Spreadsheet", type=["xlsx", "xls"], label_visibility="collapsed", key="plant_upload")
    if plant_file:
        render_selected_file_card(plant_file, "excel")
    st.markdown('</div>', unsafe_allow_html=True)
with up_col2:
    st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Invoice PDFs or ZIP</div>', unsafe_allow_html=True)
    invoice_files = st.file_uploader("Upload Invoice PDFs or ZIP", type=["pdf", "zip"], accept_multiple_files=True, label_visibility="collapsed", key="invoice_upload")
    if invoice_files:
        for _f in invoice_files[:3]:
            render_selected_file_card(_f, "pdf")
        if len(invoice_files) > 3:
            st.markdown(f'<div class="pas-file-size">+{len(invoice_files)-3} more file(s) selected</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

run = st.button("▶  Run reconciliation", use_container_width=True)

if run:
    if not plant_file or not invoice_files:
        st.warning("Please upload both the Plant workbook and invoice files/ZIP.")
        st.stop()
    try:
        with st.spinner("Reading Plant workbook..."):
            plant_df, colmap = load_plant_workbook(plant_file)
        with st.spinner("Reading invoice PDFs..."):
            invoice_records = collect_invoice_records(invoice_files)
        with st.spinner("Reconciling invoices..."):
            rows = [reconcile_invoice(inv, plant_df) for inv in invoice_records]
            all_df = pd.DataFrame(rows)
            if all_df.empty:
                st.warning("No invoice records could be extracted.")
                st.stop()
            matched_df = all_df[all_df["Match Status"] == "Matched"].copy()
            unmatched_df = all_df[all_df["Match Status"] == "Unmatched"].copy()

        total = len(all_df)
        matched = len(matched_df)
        unmatched = len(unmatched_df)
        match_pct = round((matched / total) * 100, 1) if total else 0.0

        summary_df = pd.DataFrame({
            "Metric": ["Total invoices", "Matched", "Unmatched", "Match percentage", "Run date/time"],
            "Value": [total, matched, unmatched, f"{match_pct}%", datetime.now().strftime("%d/%m/%Y %H:%M")],
        })

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M8 7V3h8l4 4v14H6V7z"/><path d="M16 3v5h5"/><path d="M9 13h6"/><path d="M9 17h4"/><path d="M4 7h2v14h12"/></svg></div><div><div class="kpi-label">Total invoices</div><div class="kpi-value">{total}</div><div class="kpi-sub">Detected records</div></div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="kpi-card kpi-matched"><div class="kpi-icon"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="9"/><path d="M8 12.5l2.7 2.7L16.5 9"/></svg></div><div><div class="kpi-label">Matched</div><div class="kpi-value">{matched}</div><div class="kpi-sub">Approved candidates</div></div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="kpi-card kpi-unmatched"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M12 3l10 18H2L12 3z"/><path d="M12 9v5"/><path d="M12 18h.01"/></svg></div><div><div class="kpi-label">Unmatched</div><div class="kpi-value">{unmatched}</div><div class="kpi-sub">Need review</div></div></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M3 20h18"/><path d="M6 16v-4"/><path d="M11 16V8"/><path d="M16 16v-6"/><path d="M19 6l-5 5-3-3-5 5"/></svg></div><div><div class="kpi-label">Match %</div><div class="kpi-value">{match_pct}%</div><div class="kpi-sub">Core KPI</div></div></div>', unsafe_allow_html=True)

        st.markdown('<div class="pas-results-title">Results</div>', unsafe_allow_html=True)
        render_unmatched_table(unmatched_df)
        
        excel_bytes = make_excel(summary_df, matched_df, unmatched_df, all_df)
        dl_left, dl_right = st.columns([1.8, 1])
        with dl_right:
            st.download_button(
                "⬇  Download Excel reconciliation",
                data=excel_bytes,
                file_name=f"PAS_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    except Exception as e:
        st.error(f"Something went wrong: {e}")
        st.exception(e)
else:
    st.info("Upload your Plant workbook and invoice PDFs/ZIP, then click Run reconciliation.")
    render_bottom_chase()
